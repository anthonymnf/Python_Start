from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Lendo pasta de trabalho da planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"]

# Pegando as referencias das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Adicionando dados e categorias nos gráficos
bar_chart = BarChart()
data = Reference(
  sheet,
  min_col=min_column + 1,
  max_col=max_column,
  min_row=min_row,
  max_row=max_row
)

categories = Reference(
  sheet,
  min_col=min_column,
  max_col=min_column,
  min_row=min_row + 1,
  max_row=max_row
)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

# Criando o gráfico 
sheet.add_chart(bar_chart, "B10")
bar_chart.title = "Vendas por fabricantes"
bar_chart.style = 2

# Salvando o Workbook
wb.save("data/bar_chart.xlsx")

